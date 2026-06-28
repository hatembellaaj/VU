import io
import re
import datetime
from typing import Union
import openpyxl
from openpyxl.utils import get_column_letter


def _try_numeric(value) -> float | None:
    """
    Tente de convertir une valeur cellule en float.
    Gère les cas :
      - int/float natifs → direct
      - strings formatées : '813 673 €', '3 086,97€', '+10.30 %', '-19.15 %'
        avec espaces insécables ( , \xa0), virgule décimale, symboles €/%/+
      - datetime → None (pas une mesure KPI)
      - None, bool → None
    """
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, datetime.datetime):
        return None
    if not isinstance(value, str):
        return None
    # Nettoyage
    s = value.strip()
    s = s.replace(' ', '').replace('\xa0', '').replace(' ', '')  # espaces insécables
    s = s.replace(' ', '')       # espaces ordinaires (séparateurs de milliers)
    s = s.replace('€', '').replace('%', '')
    s = s.replace('+', '')
    s = s.replace(',', '.')      # virgule décimale → point
    # Supprime tout sauf chiffres, point, signe moins
    s = re.sub(r'[^\d.\-]', '', s)
    if not s or s in ('.', '-', '-.'):
        return None
    try:
        return float(s)
    except ValueError:
        return None


class ExcelParser:
    """Parses Excel files (.xlsx) and extracts structured data from all sheets."""

    def parse(self, file_path_or_bytes: Union[str, bytes, io.BytesIO]) -> dict:
        """
        Parse an Excel file and extract all sheet data.

        Args:
            file_path_or_bytes: Path to file, raw bytes, or BytesIO object.

        Returns:
            {
                "sheets": {
                    sheet_name: {
                        "headers": [str, ...],
                        "rows": [[cell_value, ...], ...],
                        "numeric_cells": [
                            {"ref": "B12", "valeur": 1234.5, "sheet": "CA",
                             "row": 12, "col": 2},
                            ...
                        ]
                    }
                },
                "source": filename_or_bytes_label,
                "total_sheets": int
            }
        """
        if isinstance(file_path_or_bytes, bytes):
            source = "uploaded_bytes"
            file_obj = io.BytesIO(file_path_or_bytes)
        elif isinstance(file_path_or_bytes, io.BytesIO):
            source = getattr(file_path_or_bytes, "name", "uploaded_file")
            file_obj = file_path_or_bytes
        else:
            source = str(file_path_or_bytes)
            file_obj = file_path_or_bytes

        wb = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)

        result_sheets = {}

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            headers = []
            rows = []
            numeric_cells = []

            all_rows = list(ws.iter_rows(values_only=False))

            # Find headers: first non-empty row
            header_row_index = None
            for idx, row in enumerate(all_rows):
                row_values = [cell.value for cell in row]
                if any(v is not None and str(v).strip() != "" for v in row_values):
                    headers = [
                        str(cell.value).strip() if cell.value is not None else ""
                        for cell in row
                    ]
                    header_row_index = idx
                    break

            # Extract data rows (all rows after header)
            if header_row_index is not None:
                for row in all_rows[header_row_index + 1:]:
                    # Convertit chaque cellule : float si numérique, valeur brute sinon
                    row_values = []
                    for cell in row:
                        num = _try_numeric(cell.value)
                        row_values.append(num if num is not None else cell.value)
                    rows.append(row_values)

                    # Extract numeric cells (y compris strings formatées '813 673 €')
                    for cell in row:
                        num = _try_numeric(cell.value)
                        if num is not None:
                            col_letter = get_column_letter(cell.column)
                            ref = f"{col_letter}{cell.row}"
                            numeric_cells.append(
                                {
                                    "ref": ref,
                                    "valeur": num,
                                    "sheet": sheet_name,
                                    "row": cell.row,
                                    "col": cell.column,
                                }
                            )

                # Also check header row for numeric values
                for cell in all_rows[header_row_index]:
                    num = _try_numeric(cell.value)
                    if num is not None:
                        col_letter = get_column_letter(cell.column)
                        ref = f"{col_letter}{cell.row}"
                        numeric_cells.append(
                            {
                                "ref": ref,
                                "valeur": num,
                                "sheet": sheet_name,
                                "row": cell.row,
                                "col": cell.column,
                            }
                        )

            result_sheets[sheet_name] = {
                "headers": headers,
                "rows": rows,
                "numeric_cells": numeric_cells,
            }

        wb.close()

        return {
            "sheets": result_sheets,
            "source": source,
            "total_sheets": len(result_sheets),
        }
